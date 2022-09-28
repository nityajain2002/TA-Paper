import matplotlib.pyplot as plt
# import matplotlib.patches as mplpatches
# import matplotlib.image as mpimg
# import numpy as np
# import time
# import sys
import csv
import pandas as pd
import argparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ----- data wrangling -------------------

## set argparse (add files)
parser = argparse.ArgumentParser()
parser.add_argument('-i1', '--input_file_csv', default='TAs_Hosts.csv')
parser.add_argument('-i2', '--input_file_mlst', default='ncbi761_achtman_mlst.xlsx')
parser.add_argument('-o', '--output_file_csv', default='specific-hosts.csv')
parser.add_argument('-p', '--output_figure', default='TA_MLSTs.png')
parser.add_argument('-e', '--phylogenetic', default='USWest353 Phylogeny.xlsx')
args = parser.parse_args()
input_mlst = args.input_file_mlst
output_figure = args.output_figure
phlyogenetic = args.phylogenetic
phlyo_file = load_workbook(phlyogenetic)
phlyo = phlyo_file.active
## parse thru mslt to identify Achtman and Pasteur columns
# look into python libs for parsing excel files
# openpyxl tutorial: https://www.youtube.com/watch?v=7YS6YDQKFh0
mlst_file = load_workbook(input_mlst)
mlst = mlst_file.active
mlst_dict = {}

for row in range(2, 763):  # hard coded but have to change if more rows are added into excel
    id = get_column_letter(1)
    achtman = get_column_letter(7)
    pasteur = get_column_letter(8)
    i = mlst[id + str(row)].value.strip().split('_')
    mlst_dict[i[0]] = [mlst[achtman + str(row)].value, mlst[pasteur + str(row)].value]

mlst_file.close()
# print(mlst_dict)

## parse thru csv
ta_data = []
to_match = []
first = False
with open(args.input_file_csv, 'r') as input_csv:
    reader = csv.reader(input_csv)
    for row in reader:
        if first == False:
            first = True
            continue
        ta_data.append(row)
        # get id number + store in separate list
        t = row[2].split('_')
        to_match.append(t[0])
input_csv.close()

## add 2 columns in csv
# match values according to id
for i in range(0, len(to_match)):
    for k, v in mlst_dict.items():
        # match the id num
        if k == to_match[i]:
            a = mlst_dict[k][0]
            p = mlst_dict[k][1]
            ta_data[i][3:3] = [a, p]
            break

# is the length correct? 9934 --> 9933
# might be missing some data points
## add to tsv
# manualy create label
label = ('Domain', 'HMMER_score', 'Contig', 'Achtman', 'Pasteur',
         'Strand', 'Hit_length', 'Hit_start', 'Hit_stop',
         'Upstream_length', 'Upstream_delta', 'Upstream_start',
         'Upstream_stop', 'Downstream_length', 'Downstream_delta',
         'Downstream_start', 'Downstream_stop', 'Hit', 'Upstream', 'Downstream', 'Source')

#with open(args.output_file_csv, 'w') as output_csv:
#    writer = csv.writer(output_csv, delimiter=',',
#                        quoting=csv.QUOTE_MINIMAL)
    # add labels row manually
#    writer.writerow(label)
 #   writer.writerows(ta_data)
#output_csv.close()

# ----- generating figures -------------------

# basic figure size
fig = plt.figure(figsize=(15, 5))
# scatter plot
# ta_data[1] == HMMER_score = y axis
# ta_data[3] and [4] == Achtman and Pasteur vals
# for value in ta_data:
#    hmmer_score = value[1]
#    achtman = value[3]
#    pasteur = value[4]
#    for i in range(0,len(ta_data)):
#        panel1.scatter(achtman,hmmer_score)

# stacked bar chart
# make new lists jic there was any misalignments earlier
id = []


def ta_mlst_list(domain):
    domain_a = []
    domain_p = []
    # default list vals
    a = []
    p = []
    for line in ta_data:
        # what should our default unknown vals be? 0 for now but make it known 0 = unknown/unreported
        isA = True
        isP = True
        # set achtman
        try:
            int(line[3])
        except ValueError:
            isA = False
        if isA:
            if domain == 'overall':
                a.append(line[3])
            # disregard any unreported vals for now
            elif line[0] == domain:
                domain_a.append(line[3])
        else:
            a.append(0)
        # set pasteur
        try:
            int(line[4])
        except ValueError:
            isP = False
        if isP:
            if domain == 'overall':
                p.append(line[4])
            elif line[0] == domain:
                domain_p.append(line[4])
        else:
            p.append(0)

    # return stuff
    if domain == 'overall':
        return (a, p)
    else:
        return (domain_a, domain_p)


# counts of mlsts
def counts_mlsts(mlst_list, dict):
    #new_dict = {}
    for y in mlst_list:
        if y not in dict:
            dict[y] = 1
        else:
            dict[y] += 1
    return dict

# ydat
ydat = {}
counts_mlsts(ta_mlst_list('YdaT_toxin')[0], ydat)
#counts_mlsts(ta_mlst_list('YdaT_toxin')[1], ydat)

# parE
parE = {}
counts_mlsts(ta_mlst_list('ParE_toxin')[0], parE)
#counts_mlsts(ta_mlst_list('ParE_toxin')[1], parE)

# ccdb
ccdb = {}
counts_mlsts(ta_mlst_list('CcdB')[0], ccdb)
#counts_mlsts(ta_mlst_list('CcdB')[1], ccdb)

# pemk
pemk = {}
#counts_mlsts(ta_mlst_list('PemK_toxin')[0], pemk)
counts_mlsts(ta_mlst_list('PemK_toxin')[1], pemk)

# cpta
cpta = {}
#counts_mlsts(ta_mlst_list('Cpta_toxin')[0], cpta)
counts_mlsts(ta_mlst_list('Cpta_toxin')[1], cpta)

# gnat
gnat = {}
#counts_mlsts(ta_mlst_list('GNAT_acetyltran')[0], gnat)
counts_mlsts(ta_mlst_list('GNAT_acetyltran')[1], gnat)

# overall
overall = {}
counts_mlsts(ta_mlst_list('overall')[0], overall)
#counts_mlsts(ta_mlst_list('overall')[1], overall)
# ------ figure ---------
# both A+P
#plt.subplot(211)
#fig, stack_bar = plt.subplots()
#stack_bar.bar(ydat.keys(), ydat.values(), label='YdaT_Toxin')
#stack_bar.bar(parE.keys(), parE.values(), label='ParE_toxin')
#stack_bar.bar(pemk.keys(), pemk.values(), label='PemK_toxin')
#stack_bar.bar(cpta.keys(), cpta.values(), label='Cpta_toxin')
#stack_bar.bar(gnat.keys(), gnat.values(), label='GNAT_acetyltran')
#stack_bar.bar(ccdb.keys(), ccdb.values(), label='CcdB')

##stack_bar.set_ylabel('# of MLSTs')
#stack_bar.set_xlabel('MLST values')
#stack_bar.set_title('Pasteur MLSTs spread in NCBI TA Systems')
#stack_bar.legend()
#stack_bar.legend(bbox_to_anchor=(0,1,1,0), loc="lower left", mode="expand", ncol=2)
#fig.savefig("ncbi-ta-stacked-p.png",dpi=600)
# separate TA systems
# plt.bar(gnat.keys(),gnat.values())
# fig.savefig("GNAT_acetyltran",dpi=600)

# filter out domains
# d = {domain: [MLSTs]}
# print(len(id))
# print(len(set(id)))

# print(len(mlst_dict.keys()))
# df = pd.DataFrame(mlst_dict)
# print(ta_data)
# make the panda data frame
domain = []
ach = []
pat = []
contig = []
contig_takeout = []
takout_list = [58, 88, 224, 101, 297]
for x in ta_data:
    #print("i forgot what this did")
    #print(x)
    domain.append(x[0])
    contig.append(x[2])
    # list for reconfig_fasta
    if x[3] in takout_list and x[3] not in contig_takeout:
        contig_takeout.append(x[2])
        #print("taking this one out: ", x[2], "cuz of: ", x[3])
    if type(x[3]) != str:
        ach.append(x[3])
    else:
        ach.append(0)
    if type(x[4]) != str:
        pat.append(x[4])
    else:
        pat.append(0)

print(contig_takeout)
data = {
    'domain': domain,
    'contig': contig,
    'achtman': ach,
    'pasteur': pat,
}
df = pd.DataFrame(data)
contig_group = df.groupby('contig')
domain_group = df.groupby('domain')

groups = contig_group.groups
c_val = []
# print(len(groups.keys()))
# print(len(set(groups.keys())))
for x in domain_group.groups.keys():
    # print(x)
    if x == 'YdaT_toxin' or x == 'PemK_toxin' or x == 'ParE_toxin' or x == 'GNAT_acetyltran' or x == 'Cpta_toxin' or x == 'CcdB':
        c_val.append(list(domain_group.groups[x]))

# groups.boxplot(rot=45, fontsize=12, subplots=False)
# plt.boxplot(c_val, vert = True, labels = ['YdaT_toxin', 'PemK_toxin', 'ParE_toxin', 'GNAT', 'cpta', 'ccdb'])
# plt.title("TA systems")
# plt.ylabel("mlst values")
# plt.xlabel("domain")
# plt.savefig("domain-boxplots.png", dpi=600)

# --------------PHYLOGENETIC PROGRAM---------
phylo_a = []
phylo_p = []
for row in range(2, 353):  # hard coded but have to change if more rows are added into excel
    if type(phlyo[get_column_letter(8) + str(row)].value) != str:
        phylo_a.append(phlyo[get_column_letter(8) + str(row)].value)
    else:
        phylo_a.append(0)
    if type(phlyo[get_column_letter(9) + str(row)].value) != str:
        phylo_p.append(phlyo[get_column_letter(9) + str(row)].value)
    else:
        phylo_p.append(0)
phlyo_file.close()

def counts_mlsts_lol(mlst_list):
    new_dict = {}
    for y in mlst_list:
        if y not in new_dict:
            new_dict[y] = 1
        else:
            new_dict[y] += 1
    return new_dict

def ta_count(a, p):
    inter_ta_a = [v for v in phylo_a if v in a]
    inter_ta_p = [v for v in phylo_a if v in p]
    x = counts_mlsts_lol(inter_ta_a)
    y = counts_mlsts_lol(inter_ta_p)
    return (x, y)


def highest_mlst(mlst_dict):
    return_dict = {}
    # print("max from max", max(mlst_dict.values()))
    sorted_vals = sorted(mlst_dict.values(), reverse=True)
    # what if multiple keys have the same values?
    #print(max(ta_intersect_dict('YdaT_toxin')[0], key=lambda x: ta_intersect_dict('YdaT_toxin')[0][x]))

    for key in mlst_dict.keys():
        if mlst_dict[key] == sorted_vals[0]:
            return_dict["max"] = (key, sorted_vals[0])
        elif mlst_dict[key] == sorted_vals[1]:
            return_dict["second max"] = (key, sorted_vals[1])
        elif mlst_dict[key] == sorted_vals[2]:
            return_dict["third max"] = (key, sorted_vals[2])
    return return_dict

# check for overlaps for each ta system
def ta_intersect_dict(ta):
    # filter out the achtman + pasteur vals for each ta system
    # intersection w/ lab data
    inter_ta_a = ta_count(ta_mlst_list(ta)[0], ta_mlst_list(ta)[1])[0]
    inter_ta_p = ta_count(ta_mlst_list(ta)[0], ta_mlst_list(ta)[1])[1]
    # create stacked bar for each ta system
    #stack_bar.bar(inter_ta_a.keys(), inter_ta_a.values(), label=ta)
    return (inter_ta_a, inter_ta_p)
    #fig.savefig("ta-stacked.png",dpi=600)

# to help us distinguish a ta system for each inputted mlst val
def find_ta_system(mlst_val):
    ret_sys = set()

    # loop thru each dict to find the val
    ta_sys = set(domain)
    for i in ta_sys:
        # create the ta-intersect-dict
        temp = ta_intersect_dict(i)[0]
        # loop thru the dict to see if the given mlst val is in there
        for j in temp.keys():
            if j == mlst_val and temp[j] >= 5: # threshold of 5
                ret_sys.add(i)
    return ret_sys

def subsets(a, b):
    x = find_ta_system(a)
    y = find_ta_system(b)
    if x.issubset(y) and x != set():
        print(a, "is a subset of ", b)
    elif y.issubset(x) and y != set():
        print(b, "is a subset of", a)
    elif x == y:
        print("they're the same")
    else:
        print("inconclusive")

# given a set of trees, recursively call them
# call the find_ta_system, only account for the ones that have at least 5 MLST counts
# return set of TA systems

def phylo_tree(t):
    common = set()
    if type(t) == tuple:
        common.update(phylo_tree(t[0]))
        common.update(phylo_tree(t[1]))
    if type(t) != tuple:
        common.update(find_ta_system(t))
    common = tuple(common)
    return common

type2 = {'HicA_toxin', 'YafO_toxin', 'RelE', 'PemK_toxin', 'ParE_toxin', 'MqsR_toxin', 'Toxin_YhaV', 'HigB-like_toxin', 'Gp49', 'HigB_toxin', 'HipA_C', 'CcdB', 'YafQ_toxin'}
type3 = 'GnsAB_toxin'
type4 = 'CbtA_toxin'

def types(t):
    type = set()
    for i in t:
        if i in type2:
            type.add("2")
        if i == type3:
            type.add("3")
        if i == type4:
            type.add("4")
    return type

tree = (((((((((10,744),(167,617)),44),216),453),46),((58,88),(224,(101,297)))),((38,405),(393,69))),(117,((131,538),(135,(((73,80),95),(127,(929,(12,144))))))),(648,(354,(457,62))))

#print(phylo_tree(tree))

no_threshold = {'ParE_toxin', 'Fic', 'HigB-like_toxin', 'HD', 'AntA', 'GNAT_acetyltran', 'ANT', 'PIN', 'HipA_C', 'Polyketide_cyc2', 'CbtA_toxin', 'Cpta_toxin', 'MqsR_toxin', 'Bro-N', 'YdaT_toxin', 'YafO_toxin', 'YafQ_toxin', 'RelE', 'CcdB', 'Toxin_YhaV', 'YjhX_toxin', 'GnsAB_toxin', 'Gp49', 'DUF955', 'AbiEii', 'NTP_transf_2', 'HigB_toxin', 'HicA_toxin', 'PemK_toxin'}
threshold_5 = {'AbiEii', 'GnsAB_toxin', 'HicA_toxin', 'YafO_toxin', 'RelE', 'PIN', 'Bro-N', 'CbtA_toxin', 'Cpta_toxin', 'PemK_toxin', 'ParE_toxin', 'MqsR_toxin', 'Toxin_YhaV', 'YjhX_toxin', 'DUF955', 'HigB-like_toxin', 'Polyketide_cyc2', 'Gp49', 'HD', 'AntA', 'HigB_toxin', 'ANT', 'Fic', 'HipA_C', 'GNAT_acetyltran', 'CcdB', 'YdaT_toxin', 'YafQ_toxin'}
# only difference between the two is that no_threshold has 'NTP_transf_2'
#print(len(threshold_5))
#print(len(set(domain)))

#d1_1 = phylo_tree((12, 144))
#d2_1 = phylo_tree((929, (12, 144)))
#d3_1 = phylo_tree((127, (929, (12, 144))))
#d1_2 = phylo_tree((73, 80))
##d2_2 = phylo_tree(((73, 80), 95))
#d4 = phylo_tree(((127, (929, (12, 144))), ((73, 80), 95)))
#d5 = phylo_tree((135, ((127, (929, (12, 144))), ((73, 80), 95))))
#d1_3 = phylo_tree((131, 538))
##d6 = phylo_tree(((131, 538), (135, ((127, (929, (12, 144))), ((73, 80), 95)))))
#d7 = phylo_tree((117, ((131, 538), (135, ((127, (929, (12, 144))), ((73, 80), 95))))))

#e1 = phylo_tree((457, 62))
#e2 = phylo_tree((354, (457, 62)))
#e3 = phylo_tree((648, (354, (457, 62))))

#print(types(d1_1))
#print(types(d2_1))
#print(types(d3_1))
#print(types(d1_2))
#print(types(d2_2))
#print(types(d4))
#print(types(d5))
#print(types(d1_3))
#print(types(d6))
#print(types(d7))

#print(types(e1))
##print(types(e2))
#print(types(e3))
# stacked bar chart--
#pink = {'2', '3', '4'}

# have tp figure out the order of stacking
#plt.subplot(211)
#fig, stack_bar = plt.subplots()
#stack_bar.bar(ta_intersect_dict('YdaT_toxin')[0].keys(), ta_intersect_dict('YdaT_toxin')[0].values(), label='YdaT_toxin')
#stack_bar.bar(ta_intersect_dict('ParE_toxin')[0].keys(), ta_intersect_dict('ParE_toxin')[0].values(), label='ParE_toxin')
#stack_bar.bar(ta_intersect_dict('PemK_toxin')[0].keys(), ta_intersect_dict('PemK_toxin')[0].values(), label='PemK_toxin')
#stack_bar.bar(ta_intersect_dict('Cpta_toxin')[0].keys(), ta_intersect_dict('Cpta_toxin')[0].values(), label='Cpta_toxin')
#stack_bar.bar(ta_intersect_dict('GNAT_acetyltran')[0].keys(), ta_intersect_dict('GNAT_acetyltran')[0].values(), label='GNAT_acetyltran')
#stack_bar.bar(ta_intersect_dict('CcdB')[0].keys(), ta_intersect_dict('CcdB')[0].values(), label='CcdB')

#stack_bar.set_ylabel('# of MLSTs')
#stack_bar.set_xlabel('MLST values')
#stack_bar.set_title('Achtman Intersection with all TA Systems')
#stack_bar.legend()
#stack_bar.legend(bbox_to_anchor=(0,1,1,0), loc="lower left", mode="expand", ncol=2)
#fig.savefig("specific-ta-stacked-a.png",dpi=600)

#print(ta_intersect_dict('YdaT_toxin')[0])
#print("ydat:", highest_mlst(ta_intersect_dict('YdaT_toxin')[0]))
#print("gnat:", highest_mlst(ta_intersect_dict('GNAT_acetyltran')[0]))
#print("pemk:", highest_mlst(ta_intersect_dict('PemK_toxin')[0]))
#print("cpta:", highest_mlst(ta_intersect_dict('Cpta_toxin')[0]))
#print("ccdb:", highest_mlst(ta_intersect_dict('CcdB')[0]))
#print("pare:", highest_mlst(ta_intersect_dict('ParE_toxin')[0]))


#print("ydat:", highest_mlst(ydat))
#print("gnat:", highest_mlst(gnat))
#print("pemk:", highest_mlst(pemk))
#print("cpta:", highest_mlst(cpta))
#print("ccdb:", highest_mlst(ccdb))
#print("pare:", highest_mlst(parE))



# figures
# plt.bar(cpta_intersection.keys(),cpta_intersection.values())
# plt.xlabel("Unique MLST keys")
# plt.ylabel("Counts")
# plt.title("ccdb: NCBI and Lab Intersectional Data")
# fig.savefig("ccdb-intersection.png",dpi=600)

# panel1.set_xticks([0.5,1.5,2.5,3.5,4.5,5.5,6.5,7.5])
# panel1.set_xticklabels(['0','','6','','12','','18',''])
# panel1.set_yticks(np.arange(0,1201,200))
# add to excel

# 105 -> max ydat
# 123 -> pemk
# 67 -> pare
# 137 -> gnat
# 43 -> ccdb
# 70 -> cpta

# how to compare the two in a way that makes sense?
# 105 could just be present in the whole set more not just in ydat

# make a phlyogenitc tree for our/jay data, look for patterns in the data
# take that output and look for correlation with the ncbi data
# double checking ourselved basically
# make ncbi tree to see if they line up

# make a bar chart of all the mlst v count values

# look at second/thirdd largest to keep in mind the larger values in the data set
# look into normalizing the data? ie log scale, to show correlation better but might not be any point cuz they're lower

# 1) get TA domain numbers
# 2) MLST graphs
# 3) correlation between the two graphs
# 5) making program for our data
#    phylot, need to find one specific for tree
#    SplitsTree, Uniprot (ugene), clustalw + viewer
#    ugene builds tree + runs clustalw on it, mega is rlly slow program
#    figure out who's algorithim to use it if having trouble with ugene
# 131 might describe the generql pattern of the TA set

# 131, 122, 95, 73, 69, 648 -> possible summary of our data

# ssh mcampslabuser@dhcp-232-217.ucsc.edu
#password temp123

# cd ~/Desktop/N:C where is ugene on here?